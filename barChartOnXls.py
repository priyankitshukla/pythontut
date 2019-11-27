import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('Employee_Detail.xlsx')
sheet = wb['Sheet1']
sell_value = sheet.cell(1,1)
print(sheet.max_row)
for row in range (2,sheet.max_row+1):
    cell = sheet.cell(row,1)
    print(cell.value)
    newCol= cell.value / 100
    newCol_cell = sheet.cell(row,5)
    newCol_cell.value=newCol
value = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)
chart= BarChart()
chart.add_data(value)
sheet.add_chart(chart, 'g2')
wb.save('Employee_Detail.xlsx')