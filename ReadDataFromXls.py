import openpyxl as xl

wb = xl.load_workbook('Employee_Detail.xlsx')
sheet = wb['Sheet1']
sell_value = sheet.cell(1,1)
print(sheet.max_row)
for row in range (2,sheet.max_row+1):
    cell = sheet.cell(row,1)
    print(cell.value)
    newCol= cell.value * 5
    newCol_cell = sheet.cell(row,5)
    newCol_cell.value=newCol
wb.save('Employee_Detail2.xlsx')